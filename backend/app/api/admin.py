"""管理后台 API"""
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.police_alert import PoliceAlert
from app.models.call_record import CallRecord
from app.models.risk_supervision import RiskSupervision
from app.models.dispute_management import DisputeManagement
from app.models.display_rule import DisplayRule
from app.schemas.display_rule import DisplayRuleCreate, DisplayRuleResponse
import pandas as pd
from datetime import datetime, date
from io import BytesIO
from typing import List
from urllib.parse import quote
import json

router = APIRouter(prefix="/admin", tags=["管理后台"])


@router.get("/template")
async def download_template():
    """
    下载包含所有数据类型的多sheet模板
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    # 创建工作簿
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    # ==================== Sheet 1: 执法问题盯办 ====================
    ws1 = wb.create_sheet("执法问题盯办")
    ws1.append(["序号", "案件编号", "案件名称", "案发时间", "案件类型", "风险问题", "整改期限", "责任民警"])

    # 案件类型下拉
    dv = DataValidation(
        type="list",
        formula1='"刑事,行政,治安"',
        allow_blank=True
    )
    dv.add('E2:E1000')
    ws1.add_data_validation(dv)

    # 风险问题下拉
    risk_issues = ["案件笔录未关联", "文书未开具", "调解协议书未上传", "执法音视频未上传",
                   "涉案物品未出入库", "未结案卷未归档", "治安案件延长审批", "强制措施超期提醒"]
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(risk_issues)}"',
        allow_blank=True
    )
    dv.add('F2:F1000')
    ws1.add_data_validation(dv)

    # ==================== Sheet 2: 矛盾纠纷管理 ====================
    ws2 = wb.create_sheet("矛盾纠纷管理")
    ws2.append(["序号", "事件名称", "事件类型", "事件内容", "事发时间", "风险等级", "责任民警", "处置进度"])

    # 风险等级下拉
    dv = DataValidation(
        type="list",
        formula1='"高,中,低"',
        allow_blank=True
    )
    dv.add('F2:F1000')
    ws2.add_data_validation(dv)

    # 处置进度下拉
    dv = DataValidation(
        type="list",
        formula1='"未调解,待盯办,调解中,已调解"',
        allow_blank=True
    )
    dv.add('H2:H1000')
    ws2.add_data_validation(dv)

    # ==================== Sheet 3: 警情态势追踪（日清表） ====================
    ws3 = wb.create_sheet("警情态势追踪")
    ws3.append(["序号", "日期", "警情类型", "地点", "次数"])

    # 警情类型下拉
    dv = DataValidation(
        type="list",
        formula1='"偷盗,诈骗,涉黄,涉赌,纠纷,人身伤害"',
        allow_blank=True
    )
    dv.add('C2:C1000')
    ws3.add_data_validation(dv)

    # 次数验证
    dv = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=0,
        allow_blank=False
    )
    dv.add('E2:E1000')
    ws3.add_data_validation(dv)

    # ==================== Sheet 4: 重复报警记录（日清表） ====================
    ws4 = wb.create_sheet("重复报警记录")
    ws4.append(["序号", "日期", "报警地点", "次数"])

    # 次数验证
    dv = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1=0,
        allow_blank=False
    )
    dv.add('D2:D1000')
    ws4.add_data_validation(dv)

    # 写入内存流
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 对中文文件名进行 URL 编码
    filename = "数据导入模板.xlsx"
    encoded_filename = quote(filename)

    # 返回流式响应
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )


@router.post("/import")
async def import_data(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    导入多sheet Excel数据（一次性导入所有数据）
    """
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="只支持 Excel 文件")

    try:
        # 读取所有sheet
        excel_file = pd.ExcelFile(file.file, engine='openpyxl')

        result = {
            "risk_supervision": 0,
            "dispute_management": 0,
            "police_alert": 0,
            "call_record": 0
        }

        # ==================== 导入执法问题盯办 ====================
        if "执法问题盯办" in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name="执法问题盯办")

            for _, row in df.iterrows():
                if pd.isna(row['案件编号']):  # 跳过空行
                    continue

                case_time = pd.to_datetime(row['案发时间']) if pd.notna(row['案发时间']) else datetime.now()
                deadline = pd.to_datetime(row['整改期限']) if pd.notna(row['整改期限']) else datetime.now()
                risk_issues = str(row['风险问题']) if pd.notna(row['风险问题']) else '[]'

                stmt = sqlite_insert(RiskSupervision).values(
                    case_number=str(row['案件编号']),
                    case_name=str(row['案件名称']),
                    case_time=case_time,
                    case_type=str(row['案件类型']),
                    risk_issues=risk_issues,
                    deadline=deadline,
                    officer_name=str(row['责任民警'])
                )
                stmt = stmt.on_conflict_do_update(
                    index_elements=["case_number"],
                    set_={
                        "case_name": stmt.excluded.case_name,
                        "case_time": stmt.excluded.case_time,
                        "case_type": stmt.excluded.case_type,
                        "risk_issues": stmt.excluded.risk_issues,
                        "deadline": stmt.excluded.deadline,
                        "officer_name": stmt.excluded.officer_name,
                        "updated_at": datetime.now()
                    }
                )
                db.execute(stmt)
                result["risk_supervision"] += 1

        # ==================== 导入矛盾纠纷管理 ====================
        if "矛盾纠纷管理" in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name="矛盾纠纷管理")

            for _, row in df.iterrows():
                if pd.isna(row['事件名称']):  # 跳过空行
                    continue

                event_time = pd.to_datetime(row['事发时间']) if pd.notna(row['事发时间']) else datetime.now()

                stmt = sqlite_insert(DisputeManagement).values(
                    event_name=str(row['事件名称']),
                    event_type=str(row['事件类型']),
                    content=str(row['事件内容']),
                    event_time=event_time,
                    risk_level=str(row['风险等级']),
                    officer_name=str(row['责任民警']),
                    status=str(row['处置进度'])
                )
                stmt = stmt.on_conflict_do_update(
                    index_elements=["event_name", "event_time", "officer_name"],
                    set_={
                        "event_type": stmt.excluded.event_type,
                        "content": stmt.excluded.content,
                        "risk_level": stmt.excluded.risk_level,
                        "status": stmt.excluded.status,
                        "updated_at": datetime.now()
                    }
                )
                db.execute(stmt)
                result["dispute_management"] += 1

        # ==================== 导入警情态势追踪 ====================
        if "警情态势追踪" in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name="警情态势追踪")

            for _, row in df.iterrows():
                if pd.isna(row['日期']) or pd.isna(row['警情类型']) or pd.isna(row['地点']):  # 跳过空行
                    continue

                alert_date = pd.to_datetime(row['日期']).date()
                count = int(row['次数']) if pd.notna(row['次数']) else 0

                # 只导入次数大于0的记录
                if count > 0:
                    stmt = sqlite_insert(PoliceAlert).values(
                        alert_date=alert_date,
                        alert_type=str(row['警情类型']),
                        location=str(row['地点']),
                        count=count
                    )
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["alert_date", "alert_type", "location"],
                        set_={
                            "count": PoliceAlert.count + stmt.excluded.count
                        }
                    )
                    db.execute(stmt)
                    result["police_alert"] += 1

        # ==================== 导入重复报警记录 ====================
        if "重复报警记录" in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name="重复报警记录")

            for _, row in df.iterrows():
                if pd.isna(row['日期']) or pd.isna(row['报警地点']) or str(row['报警地点']).strip() == '':
                    continue

                call_date = pd.to_datetime(row['日期']).date()
                count = int(row['次数']) if pd.notna(row['次数']) else 0

                # 只导入次数大于0的记录
                if count > 0:
                    stmt = sqlite_insert(CallRecord).values(
                        call_date=call_date,
                        call_address=str(row['报警地点']),
                        count=count
                    )
                    stmt = stmt.on_conflict_do_update(
                        index_elements=["call_date", "call_address"],
                        set_={
                            "count": CallRecord.count + stmt.excluded.count
                        }
                    )
                    db.execute(stmt)
                    result["call_record"] += 1

        db.commit()

        return {
            "code": 200,
            "message": "导入成功",
            "data": {
                "执法问题盯办": result["risk_supervision"],
                "矛盾纠纷管理": result["dispute_management"],
                "警情态势追踪": result["police_alert"],
                "重复报警记录": result["call_record"],
                "总计": sum(result.values())
            }
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败: {str(e)}")


# ==================== 规则管理 API ====================

@router.get("/rules", response_model=dict)
async def get_rules(db: Session = Depends(get_db)):
    """
    获取所有显示规则
    """
    try:
        rules = db.query(DisplayRule).order_by(DisplayRule.priority.desc()).all()

        # 转换为响应格式
        rules_data = []
        for rule in rules:
            rule_dict = {
                "id": rule.id,
                "page_code": rule.page_code,
                "rule_type": rule.rule_type,
                "rule_name": rule.rule_name,
                "description": rule.description,
                "priority": rule.priority,
                "is_enabled": rule.is_enabled,
                "rule_config": json.loads(rule.rule_config) if rule.rule_config else {},
                "created_at": rule.created_at.isoformat() if rule.created_at else None,
                "updated_at": rule.updated_at.isoformat() if rule.updated_at else None
            }
            rules_data.append(rule_dict)

        return {
            "code": 200,
            "message": "success",
            "data": rules_data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"获取规则失败: {str(e)}")


@router.get("/rules/{rule_id}", response_model=dict)
async def get_rule(rule_id: int, db: Session = Depends(get_db)):
    """
    获取单个规则详情
    """
    rule = db.query(DisplayRule).filter(DisplayRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")

    rule_dict = {
        "id": rule.id,
        "page_code": rule.page_code,
        "rule_type": rule.rule_type,
        "rule_name": rule.rule_name,
        "description": rule.description,
        "priority": rule.priority,
        "is_enabled": rule.is_enabled,
        "rule_config": json.loads(rule.rule_config) if rule.rule_config else {},
        "created_at": rule.created_at.isoformat() if rule.created_at else None,
        "updated_at": rule.updated_at.isoformat() if rule.updated_at else None
    }

    return {
        "code": 200,
        "message": "success",
        "data": rule_dict
    }


@router.post("/rules", response_model=dict)
async def create_rule(rule_data: dict, db: Session = Depends(get_db)):
    """
    创建新规则
    """
    try:
        # 将 rule_config 转换为 JSON 字符串
        rule_config_json = json.dumps(rule_data.get("rule_config", {}), ensure_ascii=False)

        new_rule = DisplayRule(
            page_code=rule_data["page_code"],
            rule_type=rule_data["rule_type"],
            rule_name=rule_data["rule_name"],
            description=rule_data.get("description", ""),
            priority=rule_data.get("priority", 1),
            is_enabled=rule_data.get("is_enabled", 1),
            rule_config=rule_config_json
        )

        db.add(new_rule)
        db.commit()
        db.refresh(new_rule)

        return {
            "code": 200,
            "message": "创建成功",
            "data": {
                "id": new_rule.id
            }
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"创建规则失败: {str(e)}")


@router.put("/rules/{rule_id}", response_model=dict)
async def update_rule(rule_id: int, rule_data: dict, db: Session = Depends(get_db)):
    """
    更新规则
    """
    rule = db.query(DisplayRule).filter(DisplayRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")

    try:
        # 更新字段
        rule.page_code = rule_data.get("page_code", rule.page_code)
        rule.rule_type = rule_data.get("rule_type", rule.rule_type)
        rule.rule_name = rule_data.get("rule_name", rule.rule_name)
        rule.description = rule_data.get("description", rule.description)
        rule.priority = rule_data.get("priority", rule.priority)
        rule.is_enabled = rule_data.get("is_enabled", rule.is_enabled)

        # 更新 rule_config
        if "rule_config" in rule_data:
            rule.rule_config = json.dumps(rule_data["rule_config"], ensure_ascii=False)

        rule.updated_at = datetime.now()

        db.commit()

        return {
            "code": 200,
            "message": "更新成功",
            "data": {
                "id": rule.id
            }
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"更新规则失败: {str(e)}")


@router.delete("/rules/{rule_id}", response_model=dict)
async def delete_rule(rule_id: int, db: Session = Depends(get_db)):
    """
    删除规则
    """
    rule = db.query(DisplayRule).filter(DisplayRule.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="规则不存在")

    try:
        db.delete(rule)
        db.commit()

        return {
            "code": 200,
            "message": "删除成功",
            "data": None
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"删除规则失败: {str(e)}")
