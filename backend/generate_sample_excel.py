"""生成示例Excel文件用于测试导入功能

当前日期：2026-01-25
生成包含所有4个sheet的Excel文件
"""
from openpyxl import Workbook
from datetime import datetime, timedelta, date
import random

def generate_sample_excel():
    """生成示例Excel文件"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet

    today = date.today()

    # ==================== Sheet 1: 执法问题盯办 ====================
    ws1 = wb.create_sheet("执法问题盯办")
    ws1.append(["序号", "案件编号", "案件名称", "案发时间", "案件类型", "风险问题", "整改期限", "责任民警"])

    case_types = ['刑事', '行政', '治安']
    risk_issues_pool = [
        "案件笔录未关联",
        "文书未开具",
        "调解协议书未上传",
        "执法音视频未上传",
        "涉案物品未出入库",
        "未结案卷未归档",
        "治安案件延长审批",
        "强制措施超期提醒"
    ]
    officers = ['张警官', '李警官', '王警官', '赵警官', '刘警官', '陈警官']

    for i in range(15):
        case_number = f"A330903202601{str(i+1).zfill(6)}"
        case_name = f"案件{i+1}"
        case_time = (today - timedelta(days=random.randint(1, 30))).strftime('%Y-%m-%d')
        case_type = random.choice(case_types)

        # 随机1-3个风险问题，用逗号分隔
        num_issues = random.randint(1, 3)
        risk_issues = ','.join(random.sample(risk_issues_pool, num_issues))

        # 整改期限：2-15天后
        deadline = (today + timedelta(days=random.randint(2, 15))).strftime('%Y-%m-%d')
        officer_name = random.choice(officers)

        ws1.append([i+1, case_number, case_name, case_time, case_type, risk_issues, deadline, officer_name])

    # ==================== Sheet 2: 矛盾纠纷管理 ====================
    ws2 = wb.create_sheet("矛盾纠纷管理")
    ws2.append(["序号", "事件名称", "事件类型", "事件内容", "事发时间", "风险等级", "责任民警", "处置进度"])

    event_types = ['邻里矛盾', '家庭矛盾', '劳资纠纷', '物业纠纷', '其他']
    risk_levels = ['高', '中', '低']
    statuses = ['待化解', '待关注', '调解中', '已调解']

    content_templates = [
        '居民张某与李某因楼上漏水问题产生纠纷，双方情绪激动，需要及时调解处理。',
        '业主王某长期占用公共停车位，引发其他业主不满，物业协调未果。',
        '商铺租户与房东因租金上涨问题产生分歧，双方协商未果，需要调解介入。',
        '邻居因装修噪音问题产生矛盾，多次沟通无果，申请调解。',
        '小区业主因宠物饲养问题产生纠纷，需要社区介入协调。',
        '家庭成员因财产分割问题产生争议，情绪激烈，需要调解。',
        '员工与企业因工资发放问题产生劳资纠纷，申请调解。',
        '物业公司与业主因物业费收取问题产生纠纷，需要协调处理。'
    ]

    for i in range(20):
        event_name = f"纠纷事件{i+1}"
        event_type = random.choice(event_types)
        content = random.choice(content_templates)
        event_time = (today - timedelta(days=random.randint(1, 30))).strftime('%Y-%m-%d')
        risk_level = random.choice(risk_levels)
        officer_name = random.choice(officers)
        status = random.choice(statuses)

        ws2.append([i+1, event_name, event_type, content, event_time, risk_level, officer_name, status])

    # ==================== Sheet 3: 警情态势追踪 ====================
    ws3 = wb.create_sheet("警情态势追踪")
    ws3.append(["序号", "日期", "警情父类", "警情子类", "地点", "次数"])

    # 子类 → (父类, 地点列表) 配置
    alert_sub_configs = {
        '偷盗类': {
            'parent': '传统侵财',
            'locations': ['东港小区', '沈家门小区', '和平小区', '朱家尖小区', '展茅小区'],
        },
        '其它诈骗': {
            'parent': '传统侵财',
            'locations': ['东港小区', '沈家门小区', '和平小区', '朱家尖小区', '展茅小区'],
        },
        '抢夺': {
            'parent': '传统侵财',
            'locations': ['东港小区', '沈家门小区', '和平小区'],
        },
        '通讯网络诈骗': {
            'parent': '新型侵财',
            'locations': ['东港小区', '沈家门小区', '和平小区', '朱家尖小区', '展茅小区'],
        },
        '涉黄': {
            'parent': '涉黄类',
            'locations': ['东港小区', '沈家门小区', '和平小区', '朱家尖小区', '展茅小区'],
        },
        '涉赌': {
            'parent': '涉赌类',
            'locations': ['东港街道', '沈家门街道', '展茅街道', '六横镇', '桃花镇'],
        },
        '打架斗殴': {
            'parent': '人身伤害',
            'locations': ['普陀区东部', '普陀区西部', '普陀区南部', '普陀区北部', '普陀区中部'],
        },
        '家庭暴力': {
            'parent': '人身伤害',
            'locations': ['普陀区东部', '普陀区西部', '普陀区南部'],
        },
        '纠纷': {
            'parent': '纠纷类',
            'locations': ['东港社区', '沈家门社区', '展茅社区', '六横社区', '桃花社区'],
        },
    }

    # 生成最近30天的数据
    row_num = 1
    sub_type_list = list(alert_sub_configs.keys())
    for days_ago in range(30):
        alert_date = (today - timedelta(days=days_ago)).strftime('%Y-%m-%d')

        # 每天随机生成一些警情
        for sub_type in sub_type_list:
            if random.random() < 0.3:  # 30%概率
                config = alert_sub_configs[sub_type]
                location = random.choice(config['locations'])
                count = random.randint(1, 3)
                ws3.append([row_num, alert_date, config['parent'], sub_type, location, count])
                row_num += 1

    # ==================== Sheet 4: 重复报警记录 ====================
    ws4 = wb.create_sheet("重复报警记录")
    ws4.append(["序号", "日期", "报警地点", "次数"])

    locations = [
        '东港小区3号楼',
        '东港小区5号楼',
        '沈家门小区2号楼',
        '和平小区1号楼',
        '朱家尖小区4号楼',
        '展茅小区6号楼',
        '东港小区门口',
        '沈家门街道和平路128号'
    ]

    row_num = 1
    for days_ago in range(60):
        call_date = (today - timedelta(days=days_ago)).strftime('%Y-%m-%d')

        # 每天随机3-5个地点有报警
        num_locations = random.randint(3, 5)
        selected_locations = random.sample(locations, num_locations)

        for location in selected_locations:
            count = random.randint(1, 5)
            ws4.append([row_num, call_date, location, count])
            row_num += 1

    # 保存文件
    filename = "示例数据_测试导入.xlsx"
    wb.save(filename)
    print(f"✓ 已生成示例Excel文件: {filename}")
    print(f"\n文件包含:")
    print(f"  - 执法问题盯办: 15 条")
    print(f"  - 矛盾纠纷管理: 20 条")
    print(f"  - 警情态势追踪: {row_num-1} 条（最近30天）")
    print(f"  - 重复报警记录: {ws4.max_row-1} 条（最近60天）")
    print(f"\n可以使用此文件测试导入功能")


if __name__ == "__main__":
    generate_sample_excel()
